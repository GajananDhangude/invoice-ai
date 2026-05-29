import io
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    "L id", "Acc Period", "Trans Date", "Account Code", "Description",
    "Curr Code", "Trans Amount", "Dr_Cr", "Jrnal Type", "Jrnal Source",
    "Reference", "Description", "Asset Code", "Asset Indicator",
    "Asset / Item Qty", "Due Date", "Branch Analysis Code",
    "Product  Analysis Code", "ChannelAnalysisCode",
    "Sub-Channel  Analysis Code", "Underwriting Year  Analysis Code",
    "Employee Code  Analysis Code", "TDS Applicability  Analysis Code",
    "Department   Analysis Code", "Sequence Code  Analysis Code",
    "Vendor Code  Analysis Code", "Invoice Date", "From Date", "To Date",
    "Addl Date 4", "Addl Date 5", "Cheque & NEFT Number", "Invoice Number",
    "Additional Remarks", "Àdditional Remarks 2", "CREDENCE DESCRIPTION",
    "HSN/SAC NO", "Taxable on Amount", "Reverse Charge (Y/N)",
    "Reverse charge %", "Item Details (Sr.No)", "Goods/Service",
    "GST Tax Rate", "Orginal Invoice no for Dr/Cr Notes",
    "Advance Challan No ", "Addl Description 15", "Addl Description 16",
    "Addl Description 17", "Addl Description 18", "Addl Description 19",
]

# Build column map – for duplicate names, keep a list of positions
_COL_ALL: dict[str, list[int]] = {}
for idx, name in enumerate(COLUMNS, start=1):
    _COL_ALL.setdefault(name, []).append(idx)

def _col(name: str, occurrence: int = 0) -> int:
    """Return 1-based column index. Use occurrence=1 for the second 'Description'."""
    return _COL_ALL[name][occurrence]

GST_CODES = {"1120599005", "1120599007", "1120599009"}


def _excel_date(d) -> int:
    if isinstance(d, str):
        d = datetime.strptime(d, "%Y-%m-%d").date()
    elif isinstance(d, datetime):
        d = d.date()
    return (d - date(1899, 12, 30)).days


def _build_rows(inv: dict) -> list[list]:
    def blank(): return [""] * len(COLUMNS)

    def s(row, col_name, val, occurrence=0):
        row[_col(col_name, occurrence) - 1] = val

    td           = _excel_date(str(inv["invoice_date"]))
    vendor_code  = inv.get("vendor_code") or ""
    vendor_ref   = f"{vendor_code}/{inv['vendor_name'].upper()}"
    is_igst      = inv.get("igst", 0.0) > 0

    acc_period   = inv["acc_period"]
    branch       = inv.get("branch", "LXS00")
    channel      = inv.get("channel", "CC004")
    department   = inv.get("department", "SRO")
    tds_code     = inv.get("tds_code", "TD02")
    expense_acc  = inv.get("expense_account", "1310020100")
    expense_desc = inv.get("description", "VENDOR EXPENSES")
    rs_gst       = inv.get("rs_gst_number", "")
    # Col 12 "Description" = narration / billing period note
    narration    = inv.get("narration", inv.get("billing_period", ""))

    rows = []

    # ── Row 1: Expense Debit ─────────────────────────────────────────────────
    r = blank()
    s(r, "L id", "1;3;6;7")
    s(r, "Acc Period", acc_period)
    s(r, "Trans Date", td)
    s(r, "Account Code", expense_acc)
    s(r, "Description", expense_desc)              # col 5  – GL description
    s(r, "Curr Code", "INR")
    s(r, "Trans Amount", inv["taxable_amount"])
    s(r, "Dr_Cr", "D")
    s(r, "Jrnal Type", "SEXPS")
    s(r, "Reference", vendor_ref)
    s(r, "Description", narration, occurrence=1)   # col 12 – narration
    s(r, "Branch Analysis Code", branch)
    s(r, "Product  Analysis Code", "_NA")
    s(r, "ChannelAnalysisCode", channel)
    s(r, "Sub-Channel  Analysis Code", "_NA")
    s(r, "Underwriting Year  Analysis Code", "_NA")
    s(r, "Employee Code  Analysis Code", "_NA")
    s(r, "TDS Applicability  Analysis Code", tds_code)
    s(r, "Department   Analysis Code", department)
    s(r, "Vendor Code  Analysis Code", vendor_code)
    s(r, "Invoice Date", td)
    s(r, "Cheque & NEFT Number", inv["vendor_gst"])
    s(r, "Invoice Number", inv["invoice_number"])
    s(r, "HSN/SAC NO", inv.get("hsn_code", ""))
    s(r, "Orginal Invoice no for Dr/Cr Notes", rs_gst)
    rows.append(r)

    # ── Row 2: CGST Debit ────────────────────────────────────────────────────
    if not is_igst and inv.get("cgst", 0) > 0:
        r = blank()
        s(r, "L id", "6;7")
        s(r, "Acc Period", acc_period)
        s(r, "Trans Date", td)
        s(r, "Account Code", "1120599007")
        s(r, "Description", "GST input cr exps-CGST")
        s(r, "Curr Code", "INR")
        s(r, "Trans Amount", inv["cgst"])
        s(r, "Dr_Cr", "D")
        s(r, "Jrnal Type", "SEXPS")
        s(r, "Reference", vendor_ref)
        s(r, "Description", narration, occurrence=1)
        s(r, "Branch Analysis Code", branch)
        s(r, "Product  Analysis Code", "_NA")
        s(r, "ChannelAnalysisCode", channel)
        s(r, "Sub-Channel  Analysis Code", "_NA")
        s(r, "Underwriting Year  Analysis Code", "_NA")
        s(r, "Employee Code  Analysis Code", "_NA")
        s(r, "TDS Applicability  Analysis Code", "_NA")
        s(r, "Department   Analysis Code", department)
        s(r, "Vendor Code  Analysis Code", vendor_code)
        s(r, "Invoice Date", td)
        s(r, "Cheque & NEFT Number", inv["vendor_gst"])
        s(r, "Invoice Number", inv["invoice_number"])
        s(r, "HSN/SAC NO", inv.get("hsn_code", ""))
        s(r, "Taxable on Amount", inv["taxable_amount"])
        s(r, "GST Tax Rate", inv["gst_rate"] / 2)
        s(r, "Orginal Invoice no for Dr/Cr Notes", rs_gst)
        s(r, "Advance Challan No ", expense_acc)
        rows.append(r)

    # ── Row 3: SGST Debit ────────────────────────────────────────────────────
    if not is_igst and inv.get("sgst", 0) > 0:
        r = blank()
        s(r, "L id", "6;7")
        s(r, "Acc Period", acc_period)
        s(r, "Trans Date", td)
        s(r, "Account Code", "1120599005")
        s(r, "Description", "GST input cr exps-SGST")
        s(r, "Curr Code", "INR")
        s(r, "Trans Amount", inv["sgst"])
        s(r, "Dr_Cr", "D")
        s(r, "Jrnal Type", "SEXPS")
        s(r, "Reference", vendor_ref)
        s(r, "Description", narration, occurrence=1)
        s(r, "Branch Analysis Code", branch)
        s(r, "Product  Analysis Code", "_NA")
        s(r, "ChannelAnalysisCode", channel)
        s(r, "Sub-Channel  Analysis Code", "_NA")
        s(r, "Underwriting Year  Analysis Code", "_NA")
        s(r, "Employee Code  Analysis Code", "_NA")
        s(r, "TDS Applicability  Analysis Code", "_NA")
        s(r, "Department   Analysis Code", department)
        s(r, "Vendor Code  Analysis Code", vendor_code)
        s(r, "Invoice Date", td)
        s(r, "Cheque & NEFT Number", inv["vendor_gst"])
        s(r, "Invoice Number", inv["invoice_number"])
        s(r, "HSN/SAC NO", inv.get("hsn_code", ""))
        s(r, "Taxable on Amount", inv["taxable_amount"])
        s(r, "GST Tax Rate", inv["gst_rate"] / 2)
        s(r, "Orginal Invoice no for Dr/Cr Notes", rs_gst)
        s(r, "Advance Challan No ", expense_acc)
        rows.append(r)

    # ── Row 3 (alt): IGST Debit ──────────────────────────────────────────────
    elif is_igst and inv.get("igst", 0) > 0:
        r = blank()
        s(r, "L id", "6;7")
        s(r, "Acc Period", acc_period)
        s(r, "Trans Date", td)
        s(r, "Account Code", "1120599009")
        s(r, "Description", "GST input cr exps-IGST")
        s(r, "Curr Code", "INR")
        s(r, "Trans Amount", inv["igst"])
        s(r, "Dr_Cr", "D")
        s(r, "Jrnal Type", "SEXPS")
        s(r, "Reference", vendor_ref)
        s(r, "Description", narration, occurrence=1)
        s(r, "Branch Analysis Code", branch)
        s(r, "Product  Analysis Code", "_NA")
        s(r, "ChannelAnalysisCode", channel)
        s(r, "Sub-Channel  Analysis Code", "_NA")
        s(r, "Underwriting Year  Analysis Code", "_NA")
        s(r, "Employee Code  Analysis Code", "_NA")
        s(r, "TDS Applicability  Analysis Code", "_NA")
        s(r, "Department   Analysis Code", department)
        s(r, "Vendor Code  Analysis Code", vendor_code)
        s(r, "Invoice Date", td)
        s(r, "Cheque & NEFT Number", inv["vendor_gst"])
        s(r, "Invoice Number", inv["invoice_number"])
        s(r, "HSN/SAC NO", inv.get("hsn_code", ""))
        s(r, "Taxable on Amount", inv["taxable_amount"])
        s(r, "GST Tax Rate", inv["gst_rate"])
        s(r, "Orginal Invoice no for Dr/Cr Notes", rs_gst)
        s(r, "Advance Challan No ", expense_acc)
        rows.append(r)

    # ── Row 4: Vendor Credit (GST02 branch – invoice payable) ────────────────
    r = blank()
    s(r, "L id", "6;7")
    s(r, "Acc Period", acc_period)
    s(r, "Trans Date", td)
    s(r, "Account Code", vendor_code)
    s(r, "Description", inv["vendor_name"])
    s(r, "Curr Code", "INR")
    s(r, "Trans Amount", inv["net_amount"])
    s(r, "Dr_Cr", "C")
    s(r, "Jrnal Type", "SEXPS")
    s(r, "Reference", f"VT/{inv['invoice_number']}")
    s(r, "Description", narration, occurrence=1)
    s(r, "Branch Analysis Code", "GST02")
    s(r, "Product  Analysis Code", "_NA")
    s(r, "ChannelAnalysisCode", channel)
    s(r, "Sub-Channel  Analysis Code", "_NA")
    s(r, "Underwriting Year  Analysis Code", "_NA")
    s(r, "Employee Code  Analysis Code", "_NA")
    s(r, "TDS Applicability  Analysis Code", "_NA")
    s(r, "Department   Analysis Code", department)
    s(r, "Vendor Code  Analysis Code", vendor_code)
    s(r, "Invoice Date", td)
    s(r, "Cheque & NEFT Number", inv["vendor_gst"])
    s(r, "Invoice Number", inv["invoice_number"])
    rows.append(r)

    return rows


def build_journal_excel(invoices: list[dict]) -> bytes:
    """
    Builds journal entry Excel entirely in memory.
    No template file needed — header written from scratch.
    Returns bytes ready for StreamingResponse.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "upload format"

    # Styles
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F3864")
    data_font   = Font(name="Arial", size=9)
    dr_fill     = PatternFill("solid", start_color="E8F4FD")   # blue   – Debit
    cr_fill     = PatternFill("solid", start_color="FFF2CC")   # yellow – Credit
    gst_fill    = PatternFill("solid", start_color="E2EFDA")   # green  – GST
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center")

    # ── Header row ───────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = border
        cell.alignment = center
    ws.row_dimensions[1].height = 30

    # ── Column widths ────────────────────────────────────────────────────────
    widths = {
        "L id": 8, "Acc Period": 10, "Trans Date": 14, "Account Code": 16,
        "Description": 30, "Curr Code": 8, "Trans Amount": 14, "Dr_Cr": 6,
        "Jrnal Type": 10, "Reference": 32, "Invoice Number": 20,
        "Cheque & NEFT Number": 22, "HSN/SAC NO": 12,
        "Taxable on Amount": 16, "GST Tax Rate": 10,
        "Vendor Code  Analysis Code": 16, "Invoice Date": 14,
        "Branch Analysis Code": 16, "ChannelAnalysisCode": 14,
        "Orginal Invoice no for Dr/Cr Notes": 24, "Advance Challan No ": 18,
        "Department   Analysis Code": 14, "TDS Applicability  Analysis Code": 14,
    }
    for col_name, width in widths.items():
        # Use first occurrence for duplicates
        ws.column_dimensions[get_column_letter(_col(col_name))].width = width
        if col_name == "Description":
            ws.column_dimensions[get_column_letter(_col(col_name, 1))].width = width

    ws.freeze_panes = "A2"

    # ── Data rows ────────────────────────────────────────────────────────────
    row_idx = 2
    for inv in invoices:
        for jrow in _build_rows(inv):
            account_code = str(jrow[_col("Account Code") - 1])
            dr_cr        = jrow[_col("Dr_Cr") - 1]

            if account_code in GST_CODES:
                fill = gst_fill
            elif dr_cr == "C":
                fill = cr_fill
            else:
                fill = dr_fill

            for col_idx, val in enumerate(jrow, start=1):
                cell           = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font      = data_font
                cell.fill      = fill
                cell.border    = border
                cell.alignment = left

                col_name = COLUMNS[col_idx - 1]
                if col_name in ("Trans Date", "Invoice Date") and isinstance(val, int) and val > 0:
                    cell.number_format = "DD-MMM-YYYY"
                if col_name in ("Trans Amount", "Taxable on Amount"):
                    cell.number_format = "#,##0.00"
                if col_name == "GST Tax Rate" and isinstance(val, (int, float)) and val:
                    cell.number_format = '0.00"%"'

            row_idx += 1
        row_idx += 1  # blank separator between invoices

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()